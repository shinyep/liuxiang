import os
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from datetime import datetime
from api.models import Customer, Shipment, ProductionTimeSlot

class Command(BaseCommand):
    help = 'Imports shipment data from an XLSX file into Shipment and ProductionTimeSlot models'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='The path to the XLSX file to import.')
        parser.add_argument(
            '--sheet',
            type=str,
            help='Name of the sheet to import. Defaults to the first active sheet.'
        )

    def handle(self, *args, **options):
        file_path = options['file_path']
        sheet_name = options.get('sheet')

        if not os.path.exists(file_path):
            raise CommandError(f"File not found at path: {file_path}")

        try:
            workbook = load_workbook(filename=file_path, data_only=True)
            sheet = workbook[sheet_name] if sheet_name else workbook.active
        except Exception as e:
            raise CommandError(f"Error loading XLSX file: {e}")

        # Define expected header mapping (0-indexed) - adjust based on actual file
        # This is an assumed mapping based on the provided images.
        # A=0 (Area Code), B=1 (Customer Name), C=2 (Shipping Address),
        # D=3 (Material No), E=4 (Material Desc), F=5 (Outer Box Date)
        # G=6 (Production Date Code), H=7 (Batch), I=8 (Spec/Quantity per slot), J=9 (Packaging Line)
        # K=10 (Time Slot 1), L=11 (Time Slot 2), ... up to 10 slots (T=19)
        COL_MAP = {
            'area_code': 0,
            'customer_name': 1,
            'shipping_address': 2,
            'material_number': 3,
            'material_description': 4,
            'outer_box_date': 5, # Excel F
            'production_date_code': 6, # Excel G
            'batch_number': 7, # Excel H - assuming this is batch like "YT20241205" in screenshot F, but data like 12.4 in screenshot G.
                               # User said this is "批次" and "YT20241205" is "生产日期码". The new screenshot matches excel G with data like "YT20241205", H is "批次" with 12.4
                               # So, F = outer_box_date (date type), G = production_date_code (string), H = batch_number (string/float), J=packaging_line (original excel H)
                               # I = specification (string/int, used as quantity for slots) (original excel I)
            'specification_quantity': 8, # Excel I
            'packaging_line': 9, # Excel J
            'time_slot_start_col': 10,  # Columns K to T (10 slots)
            'time_slot_cols': 10 # Number of time slot columns
        }
        # The column definition for "batch_number" in excel is G in the provided image (e.g., YT20241203)
        # The column "规格" is H (e.g. 168)
        # The column "包装线" is I (e.g. A)
        # The "起止时间1" starts at J
        # The new Excel screenshot has:
        # A: 区域码, B: 客户名, C: 送货地址, D: 物料编号, E: 物料描述
        # F: 外箱日期, G: 生产日期码 (eg YT20241205), H: 批次 (eg 12.3), I: 规格 (eg 168), J: 包装线 (eg A)
        # K to T: 起止时间 1 to 10
        
        # Corrected mapping based on further clarification and "起止时间里有字母F" screenshot
        COL_MAP_REVISED = {
            'area_code': 0,             # A
            'customer_name': 1,         # B
            'shipping_address': 2,      # C
            'material_number': 3,       # D
            'material_description': 4,  # E
            'outer_box_date': 5,        # F
            'production_date_code': 6,  # G (e.g., YT20241205)
            'batch_number': 7,          # H (e.g., 12.4)
            'specification_quantity': 8,# I (e.g., 168, this will be the quantity for slots)
            'packaging_line': 9,        # J (e.g., A)
            'time_slot_start_col': 10,  # K (起止时间1)
            'num_time_slots': 10        # Up to 10 time slots (K to T)
        }
        
        current_map = COL_MAP_REVISED

        header = [cell.value for cell in sheet[1]]
        self.stdout.write(self.style.SUCCESS(f"Importing data from sheet: {sheet.title}"))
        self.stdout.write(f"Header: {header}")

        imported_count = 0
        failed_rows = []

        with transaction.atomic():
            # Start from row 2 (assuming row 1 is header)
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    area_code = str(row[current_map['area_code']]).strip() if row[current_map['area_code']] else None
                    customer_name = str(row[current_map['customer_name']]).strip() if row[current_map['customer_name']] else None
                    shipping_address = str(row[current_map['shipping_address']]).strip() if row[current_map['shipping_address']] else ''
                    
                    if not area_code or not customer_name:
                        self.stderr.write(self.style.WARNING(f"Skipping row {row_num}: Missing area_code or customer_name."))
                        failed_rows.append({'row': row_num, 'reason': 'Missing area_code or customer_name'})
                        continue

                    # Get or create customer
                    customer, created = Customer.objects.get_or_create(
                        area_code=area_code,
                        defaults={'name': customer_name, 'shipping_address': shipping_address}
                    )
                    if not created and customer.name != customer_name: # Update if name changed
                        customer.name = customer_name
                        customer.shipping_address = shipping_address # Also update address if customer exists
                        customer.save()
                    elif not created and customer.shipping_address != shipping_address:
                         customer.shipping_address = shipping_address
                         customer.save()


                    outer_box_date_val = row[current_map['outer_box_date']]
                    if isinstance(outer_box_date_val, datetime):
                        outer_box_date = outer_box_date_val.date()
                    elif isinstance(outer_box_date_val, str):
                        try:
                            outer_box_date = datetime.strptime(outer_box_date_val, '%Y-%m-%d').date()
                        except ValueError:
                             outer_box_date = None # Or handle as error
                    else: # Might be int from Excel date
                        try:
                            # Assuming standard excel date (number of days since 1899-12-30)
                            from openpyxl.utils.datetime import from_excel
                            outer_box_date = from_excel(outer_box_date_val).date()
                        except:
                            outer_box_date = None

                    if not outer_box_date:
                        self.stderr.write(self.style.WARNING(f"Skipping row {row_num} for customer {customer_name}: Invalid or missing outer_box_date."))
                        failed_rows.append({'row': row_num, 'reason': f'Invalid outer_box_date for {customer_name}'})
                        continue
                        
                    shipment_data = {
                        'customer': customer,
                        'material_number': str(row[current_map['material_number']]).strip() if row[current_map['material_number']] else '',
                        'material_description': str(row[current_map['material_description']]).strip() if row[current_map['material_description']] else '',
                        'outer_box_date': outer_box_date,
                        'production_date_code': str(row[current_map['production_date_code']]).strip() if row[current_map['production_date_code']] else '',
                        'batch_number': str(row[current_map['batch_number']]).strip() if row[current_map['batch_number']] else '',
                        'specification': str(row[current_map['specification_quantity']]).strip() if row[current_map['specification_quantity']] else '', # this is the overall spec/qty for the shipment row
                        'packaging_line': str(row[current_map['packaging_line']]).strip() if row[current_map['packaging_line']] else '',
                    }
                    
                    # Use 'specification_quantity' as the quantity for ALL time slots in this row.
                    try:
                        slot_quantity_val = row[current_map['specification_quantity']]
                        slot_quantity = int(float(str(slot_quantity_val))) if slot_quantity_val is not None else 0
                    except (ValueError, TypeError):
                        self.stderr.write(self.style.WARNING(f"Row {row_num}: Invalid quantity value '{slot_quantity_val}', using 0."))
                        slot_quantity = 0


                    shipment = Shipment.objects.create(**shipment_data)

                    # Import ProductionTimeSlots
                    for i in range(current_map['num_time_slots']):
                        col_idx = current_map['time_slot_start_col'] + i
                        if col_idx >= len(row) or row[col_idx] is None: # Check if column exists / has data
                            continue 
                        
                        time_slot_raw = str(row[col_idx]).strip()
                        if not time_slot_raw: # Skip if cell is empty
                            continue

                        start_time_str = ''
                        end_time_str = ''

                        if time_slot_raw.upper() == 'F':
                            start_time_str = 'F'
                            end_time_str = '' # As per user request to display only "F"
                        else:
                            # Try to parse HHMM-HHMM or HHMM
                            parts = time_slot_raw.split('-')
                            if len(parts) == 2:
                                start_time_str = parts[0].strip()
                                end_time_str = parts[1].strip()
                                # Basic validation for HHMM format (can be enhanced)
                                if not (len(start_time_str) == 4 and start_time_str.isdigit() and \
                                        len(end_time_str) == 4 and end_time_str.isdigit()):
                                    self.stderr.write(self.style.WARNING(f"Row {row_num}, Slot {i+1}: Invalid time format '{time_slot_raw}', skipping slot."))
                                    continue 
                            elif len(parts) == 1 and len(time_slot_raw) == 4 and time_slot_raw.isdigit(): # Single HHMM
                                start_time_str = time_slot_raw
                                end_time_str = '' # Or perhaps duplicate start_time_str if it means a point in time? For now, empty.
                            else: # Unrecognized format
                                self.stderr.write(self.style.WARNING(f"Row {row_num}, Slot {i+1}: Unrecognized time format '{time_slot_raw}', skipping slot."))
                                continue
                        
                        if start_time_str: # Only create if we have a start time (or "F")
                            ProductionTimeSlot.objects.create(
                                shipment=shipment,
                                start_time_str=start_time_str,
                                end_time_str=end_time_str,
                                quantity=slot_quantity # Use the quantity from the 'specification_quantity' column for this slot
                            )
                    imported_count += 1
                except Exception as e:
                    self.stderr.write(self.style.ERROR(f"Error processing row {row_num}: {e} - Data: {row}"))
                    failed_rows.append({'row': row_num, 'reason': str(e)})
        
        if failed_rows:
            self.stdout.write(self.style.WARNING(f"\nImport completed with some errors. {len(failed_rows)} rows failed:"))
            for fail in failed_rows:
                self.stdout.write(self.style.WARNING(f"  Row {fail['row']}: {fail['reason']}"))
        
        self.stdout.write(self.style.SUCCESS(f"\nSuccessfully imported {imported_count} shipment records."))
